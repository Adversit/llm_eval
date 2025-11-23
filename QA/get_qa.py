import pandas as pd
import json
import requests
import os
from openpyxl import load_workbook
from QA.config import API_KEY, MODEL, API_BASE_URL, DEFAULT_TIMEOUT, DEFAULT_MAX_TOKENS, DEFAULT_TEMPERATURE

def generate_qa_pairs(title, content, num_pairs=5):
    """
    使用大模型基于标题和内容生成问答对

    参数:
        title (str): 文章标题
        content (str): 标题对应的内容
        num_pairs (int): 希望生成的问答对数量

    返回:
        list: 包含生成的问答对列表
    """
    try:
        # 准备提示词
        prompt = f"""
请基于以下标题和内容生成{num_pairs}个高质量的问答对。问题应该是关于内容的重要信息，答案必须能够从提供的内容中找到依据。

标题: {title}

内容:
{content}

要求：
1. 生成{num_pairs}个问答对
2. 问题应该有认知深度，答案要有实用价值
3. 答案必须严格基于提供的内容，不要编造信息
4. 问题应该涵盖内容中的关键点，而不是细枝末节
5. 答案应该简洁明了，直接回答问题

请以JSON格式返回，格式为：
{{
    "qa_pairs": [
        {{
            "question": "问题1",
            "answer": "答案1"
        }},
        ...
    ]
}}
"""
        
        # 调用API
        url = f"{API_BASE_URL}/v1/chat/completions"

        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": MODEL,
            "messages": [
                {
                    "role": "system",
                    "content": "你是一个专业的问答对生成专家，擅长从文本中提取关键信息并创建有价值的问答对。请确保生成的问答对准确反映原文内容，并且具有教育价值。"
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "stream": False,
            "max_tokens": DEFAULT_MAX_TOKENS,
            "temperature": DEFAULT_TEMPERATURE,
            "response_format": {"type": "json_object"}
        }

        # 发送请求
        response = requests.request("POST", url, json=payload, headers=headers, timeout=DEFAULT_TIMEOUT)
        
        if response.status_code == 200:
            response_data = response.json()
            content = response_data['choices'][0]['message']['content']
            
            # 解析JSON响应
            try:
                result = json.loads(content)
                
                # 确保结果包含预期的字段
                if 'qa_pairs' not in result:
                    result['qa_pairs'] = []
                
                return {
                    'success': True,
                    'qa_pairs': result['qa_pairs']
                }
                
            except json.JSONDecodeError:
                return {
                    'success': False,
                    'error': '无法解析API响应为JSON',
                    'raw_response': content,
                    'qa_pairs': []
                }
        else:
            return {
                'success': False,
                'error': f'API调用失败: {response.status_code} - {response.text}',
                'qa_pairs': []
            }

    except requests.Timeout:
        return {
            'success': False,
            'error': f'API请求超时（超过{DEFAULT_TIMEOUT}秒）',
            'qa_pairs': []
        }
    except requests.RequestException as e:
        return {
            'success': False,
            'error': f'网络请求错误: {str(e)}',
            'qa_pairs': []
        }
    except Exception as e:
        return {
            'success': False,
            'error': f'生成问答对时出错: {str(e)}',
            'qa_pairs': []
        }

def process_excel_and_generate_qa(excel_path, output_path=None, num_pairs_per_section=5, use_suggested_count=False):
    """
    处理Excel文件并为每个标题-内容对生成问答对
    
    参数:
        excel_path (str): 输入Excel文件路径
        output_path (str, optional): 输出Excel文件路径，默认添加"_QA"后缀
        num_pairs_per_section (int): 每个部分生成的问答对数量，默认为5
        use_suggested_count (bool): 是否使用建议的问答对数量，默认为False
    
    返回:
        bool: 操作是否成功
    """
    try:
        # 设置默认输出路径
        if output_path is None:
            base_name, ext = os.path.splitext(excel_path)
            output_path = f"{base_name}_QA{ext}"
        
        print(f"处理Excel文件: {excel_path}")
        if use_suggested_count:
            print("将使用建议的问答对数量")
        else:
            print(f"每个部分将生成 {num_pairs_per_section} 个问答对")
        
        # 读取Excel文件
        df = pd.read_excel(excel_path)
        
        # 确保必要的列存在
        required_columns = ['标题', '内容']
        if not all(col in df.columns for col in required_columns):
            print(f"错误: Excel文件必须包含以下列: {', '.join(required_columns)}")
            return False
        
        # 检查是否有建议问答对数量列
        has_suggested_count = '建议问答对数量' in df.columns
        
        # 如果要使用建议数量但没有该列，发出警告
        if use_suggested_count and not has_suggested_count:
            print("警告: 未找到'建议问答对数量'列，将使用默认数量")
            use_suggested_count = False
        
        # 创建新数据框来存储结果
        result_data = []
        
        # 遍历每个标题和内容对
        total_rows = len(df)
        for index, row in df.iterrows():
            title = row['标题']
            content = row['内容']
            
            # 确定要生成的问答对数量
            if use_suggested_count and has_suggested_count:
                suggested_count = row.get('建议问答对数量', num_pairs_per_section)
                # 确保建议数量是有效的整数且大于0
                if pd.isna(suggested_count) or not isinstance(suggested_count, (int, float)):
                    pairs_to_generate = num_pairs_per_section
                else:
                    suggested_count_int = int(suggested_count)
                    # 验证数值范围（必须大于0，建议不超过20）
                    if suggested_count_int <= 0:
                        print(f"  警告: 建议问答对数量({suggested_count_int})无效，使用默认值{num_pairs_per_section}")
                        pairs_to_generate = num_pairs_per_section
                    elif suggested_count_int > 20:
                        print(f"  警告: 建议问答对数量({suggested_count_int})过大，限制为20")
                        pairs_to_generate = 20
                    else:
                        pairs_to_generate = suggested_count_int
            else:
                pairs_to_generate = num_pairs_per_section
            
            print(f"正在处理 [{index+1}/{total_rows}]: {title} (生成 {pairs_to_generate} 个问答对)")
            
            # 如果内容为空，则跳过
            if pd.isna(content) or content.strip() == '':
                print(f"  跳过（内容为空）: {title}")
                continue
                
            # 生成问答对
            qa_result = generate_qa_pairs(title, content, pairs_to_generate)
            
            if qa_result['success']:
                # 为每个生成的问答对创建一行
                for i, qa_pair in enumerate(qa_result['qa_pairs']):
                    result_data.append({
                        '标题': title,
                        '问题编号': f"{index+1}-{i+1}",
                        '问题': qa_pair['question'],
                        '答案': qa_pair['answer'],
                        '内容': content  # 添加原始内容列
                    })
                print(f"  成功生成 {len(qa_result['qa_pairs'])} 个问答对")
            else:
                print(f"  处理失败: {qa_result.get('error', '未知错误')}")
        
        # 创建DataFrame并保存为Excel
        if result_data:
            result_df = pd.DataFrame(result_data)
            
            # 调整Excel列宽并保存
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False)
                
                # 获取工作表
                worksheet = writer.sheets['Sheet1']
                
                # 设置列宽
                column_widths = {
                    'A': 40,  # 标题
                    'B': 10,  # 问题编号
                    'C': 50,  # 问题
                    'D': 80,  # 答案
                    'E': 80,  # 内容
                }
                
                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
                
                # 设置自动换行
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = worksheet.cell(1, 1).alignment.copy(wrapText=True)
            
            print(f"已成功生成 {len(result_data)} 个问答对并保存到Excel文件: {output_path}")
            return True
        else:
            print("没有生成任何问答对，未创建输出文件")
            return False
            
    except Exception as e:
        print(f"处理Excel文件时出错: {str(e)}")
        return False

if __name__ == "__main__":
    # 设置输入文件路径
    input_excel = "标题内容提取结果_evaluated_filtered.xlsx"
    
    # 处理Excel文件并生成问答对
    process_excel_and_generate_qa(input_excel, num_pairs_per_section=3)