import pandas as pd
import json
import requests
import os
from openpyxl import load_workbook
from config import API_KEY, MODEL, API_BASE_URL

def evaluate_content_quality(title, content, include_reason=True, suggest_qa_count=False):
    """
    使用大模型评估内容的质量，为信息密度和信息质量打分

    参数:
        title (str): 文章标题
        content (str): 标题对应的内容
        include_reason (bool): 是否包含评估理由，默认为True
        suggest_qa_count (bool): 是否建议问答对数量，默认为False

    返回:
        dict: 包含评估结果的字典
    """
    # 使用统一的API配置
    api_key = API_KEY
    model = MODEL
    api_base_url = API_BASE_URL
    
    try:
        # 准备提示词
        if include_reason and suggest_qa_count:
            prompt = f"""
请评估以下标题和内容的质量，并为其信息密度和信息质量打分。

标题: {title}

内容:
{content}

评估要求：
1. 信息密度评分(1-10分)：评估内容中实质性信息的丰富程度，高分表示内容包含大量有价值的信息，低分表示内容空洞或仅包含目录、标题等
2. 信息质量评分(1-10分)：评估内容的专业性、准确性和实用价值
3. 是否值得处理(true/false)：根据内容质量判断是否值得进一步处理生成问答对
4. 建议问答对数量(1-10个)：根据内容质量和丰富程度，建议从该内容中生成多少个问答对最为合适
5. 评估理由：简要说明评分和判断的依据

请以JSON格式返回，格式为：
{{
    "density_score": 评分(1-10),
    "quality_score": 评分(1-10),
    "worth_processing": true或false,
    "suggested_qa_count": 建议数量(1-10),
    "evaluation_reason": "评估理由"
}}
"""
        elif include_reason and not suggest_qa_count:
            prompt = f"""
请评估以下标题和内容的质量，并为其信息密度和信息质量打分。

标题: {title}

内容:
{content}

评估要求：
1. 信息密度评分(1-10分)：评估内容中实质性信息的丰富程度，高分表示内容包含大量有价值的信息，低分表示内容空洞或仅包含目录、标题等
2. 信息质量评分(1-10分)：评估内容的专业性、准确性和实用价值
3. 是否值得处理(true/false)：根据内容质量判断是否值得进一步处理生成问答对
4. 评估理由：简要说明评分和判断的依据

请以JSON格式返回，格式为：
{{
    "density_score": 评分(1-10),
    "quality_score": 评分(1-10),
    "worth_processing": true或false,
    "evaluation_reason": "评估理由"
}}
"""
        elif not include_reason and suggest_qa_count:
            prompt = f"""
请评估以下标题和内容的质量，并为其信息密度和信息质量打分。

标题: {title}

内容:
{content}

评估要求：
1. 信息密度评分(1-10分)：评估内容中实质性信息的丰富程度，高分表示内容包含大量有价值的信息，低分表示内容空洞或仅包含目录、标题等
2. 信息质量评分(1-10分)：评估内容的专业性、准确性和实用价值
3. 是否值得处理(true/false)：根据内容质量判断是否值得进一步处理生成问答对
4. 建议问答对数量(1-8个)：根据目标文档中的信息数量，建议从该内容中生成多少个问答对最为合适，采用保守原则。

请以JSON格式返回，格式为：
{{
    "density_score": 评分(1-10),
    "quality_score": 评分(1-10),
    "worth_processing": true或false,
    "suggested_qa_count": 建议数量(1-8)
}}
"""
        else:  # not include_reason and not suggest_qa_count
            prompt = f"""
请评估以下标题和内容的质量，并为其信息密度和信息质量打分。

标题: {title}

内容:
{content}

评估要求：
1. 信息密度评分(1-10分)：评估内容中实质性信息的丰富程度，高分表示内容包含大量有价值的信息，低分表示内容空洞或仅包含目录、标题等
2. 信息质量评分(1-10分)：评估内容的专业性、准确性和实用价值
3. 是否值得处理(true/false)：根据内容质量判断是否值得进一步处理生成问答对

请以JSON格式返回，格式为：
{{
    "density_score": 评分(1-10),
    "quality_score": 评分(1-10),
    "worth_processing": true或false
}}
"""
        
        # 调用API
        url = f"{api_base_url}/v1/chat/completions"
        
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": model,
            "messages": [
                {
                    "role": "system", 
                    "content": "你是一个专业的内容质量评估专家，擅长判断文本的信息密度和质量。请确保评估客观公正，并给出合理的分数和判断。"
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "stream": False,
            "max_tokens": 1000 if include_reason else 500,
            "temperature": 0.2,
            "response_format": {"type": "json_object"}
        }
        
        # 发送请求
        response = requests.request("POST", url, json=payload, headers=headers, timeout=60)
        
        if response.status_code == 200:
            response_data = response.json()
            content = response_data['choices'][0]['message']['content']
            
            # 解析JSON响应
            try:
                result = json.loads(content)
                
                # 确保结果包含预期的字段
                expected_fields = ['density_score', 'quality_score', 'worth_processing']
                if include_reason:
                    expected_fields.append('evaluation_reason')
                if suggest_qa_count:
                    expected_fields.append('suggested_qa_count')
                    
                for field in expected_fields:
                    if field not in result:
                        if field in ['density_score', 'quality_score']:
                            result[field] = 0
                        elif field == 'worth_processing':
                            result[field] = False
                        elif field == 'suggested_qa_count':
                            result[field] = 3  # 默认建议3个问答对
                        else:
                            result[field] = ""
                
                # 如果不包含评估理由但API返回了，则删除
                if not include_reason and 'evaluation_reason' in result:
                    del result['evaluation_reason']
                # 如果需要评估理由但API没有返回，则添加空字符串
                elif include_reason and 'evaluation_reason' not in result:
                    result['evaluation_reason'] = ""
                
                # 如果不需要建议问答对数量但API返回了，则删除
                if not suggest_qa_count and 'suggested_qa_count' in result:
                    del result['suggested_qa_count']
                # 如果需要建议问答对数量但API没有返回，则添加默认值
                elif suggest_qa_count and 'suggested_qa_count' not in result:
                    result['suggested_qa_count'] = 3
                
                # 确保建议的问答对数量在1-10之间
                if suggest_qa_count and 'suggested_qa_count' in result:
                    count = result['suggested_qa_count']
                    if not isinstance(count, int) or count < 1:
                        result['suggested_qa_count'] = 1
                    elif count > 10:
                        result['suggested_qa_count'] = 10
                
                return {
                    'success': True,
                    'evaluation': result
                }
                
            except json.JSONDecodeError:
                default_evaluation = {
                    'density_score': 0,
                    'quality_score': 0,
                    'worth_processing': False
                }
                if include_reason:
                    default_evaluation['evaluation_reason'] = '评估失败'
                if suggest_qa_count:
                    default_evaluation['suggested_qa_count'] = 3
                    
                return {
                    'success': False,
                    'error': '无法解析API响应为JSON',
                    'raw_response': content,
                    'evaluation': default_evaluation
                }
        else:
            default_evaluation = {
                'density_score': 0,
                'quality_score': 0,
                'worth_processing': False
            }
            if include_reason:
                default_evaluation['evaluation_reason'] = 'API调用失败'
            if suggest_qa_count:
                default_evaluation['suggested_qa_count'] = 3
                
            return {
                'success': False,
                'error': f'API调用失败: {response.status_code} - {response.text}',
                'evaluation': default_evaluation
            }
            
    except Exception as e:
        default_evaluation = {
            'density_score': 0,
            'quality_score': 0,
            'worth_processing': False
        }
        if include_reason:
            default_evaluation['evaluation_reason'] = f'处理错误: {str(e)}'
        if suggest_qa_count:
            default_evaluation['suggested_qa_count'] = 3
            
        return {
            'success': False,
            'error': f'评估内容质量时出错: {str(e)}',
            'evaluation': default_evaluation
        }

def process_excel_and_evaluate(input_excel, output_excel=None, min_density_score=5, min_quality_score=5, include_reason=True, suggest_qa_count=False):
    """
    处理Excel文件并评估每个标题-内容对的质量
    
    参数:
        input_excel (str): 输入Excel文件路径
        output_excel (str, optional): 输出Excel文件路径
        min_density_score (int): 最低信息密度分数阈值
        min_quality_score (int): 最低信息质量分数阈值
        include_reason (bool): 是否包含评估理由，默认为True
        suggest_qa_count (bool): 是否建议问答对数量，默认为False
    
    返回:
        bool: 操作是否成功
    """
    try:
        # 设置默认输出路径
        if output_excel is None:
            base_name, ext = os.path.splitext(input_excel)
            output_excel = f"{base_name}_evaluated{ext}"
        
        # 读取Excel文件
        df = pd.read_excel(input_excel)
        
        # 确保必要的列存在
        required_columns = ['标题', '内容']
        if not all(col in df.columns for col in required_columns):
            print(f"错误: Excel文件必须包含以下列: {', '.join(required_columns)}")
            return False
        
        # 添加评估结果列
        df['信息密度评分'] = 0
        df['信息质量评分'] = 0
        df['是否值得处理'] = False
        if include_reason:
            df['评估理由'] = ''
        if suggest_qa_count:
            df['建议问答对数量'] = 3  # 默认值
        
        # 遍历每个标题和内容对
        total_rows = len(df)
        for index, row in df.iterrows():
            title = row['标题']
            content = row['内容']
            
            print(f"正在评估 [{index+1}/{total_rows}]: {title}")
            
            # 如果内容为空，则跳过
            if pd.isna(content) or content.strip() == '':
                print(f"  跳过（内容为空）: {title}")
                if include_reason:
                    df.at[index, '评估理由'] = '内容为空'
                continue
                
            # 评估内容质量
            eval_result = evaluate_content_quality(title, content, include_reason, suggest_qa_count)
            
            if eval_result['success']:
                evaluation = eval_result['evaluation']
                
                # 更新DataFrame
                df.at[index, '信息密度评分'] = evaluation['density_score']
                df.at[index, '信息质量评分'] = evaluation['quality_score']
                df.at[index, '是否值得处理'] = evaluation['worth_processing']
                if include_reason and 'evaluation_reason' in evaluation:
                    df.at[index, '评估理由'] = evaluation['evaluation_reason']
                if suggest_qa_count and 'suggested_qa_count' in evaluation:
                    df.at[index, '建议问答对数量'] = evaluation['suggested_qa_count']
                
                qa_count_info = f", 建议问答对数量={evaluation.get('suggested_qa_count', '未提供')}" if suggest_qa_count else ""
                print(f"  评估结果: 密度={evaluation['density_score']}, 质量={evaluation['quality_score']}, 值得处理={evaluation['worth_processing']}{qa_count_info}")
            else:
                print(f"  评估失败: {eval_result.get('error', '未知错误')}")
                if include_reason:
                    df.at[index, '评估理由'] = f"评估失败: {eval_result.get('error', '未知错误')}"
        
        # 创建一个筛选后的DataFrame
        filtered_df = df[
            (df['信息密度评分'] >= min_density_score) & 
            (df['信息质量评分'] >= min_quality_score) & 
            (df['是否值得处理'] == True)
        ].copy()
        
        # 保存完整评估结果
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='全部评估结果', index=False)
            filtered_df.to_excel(writer, sheet_name='筛选后结果', index=False)
            
            # 设置列宽
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                column_widths = {
                    'A': 10,  # 标题行号列（如果有）
                    'B': 40,  # 标题列
                    'C': 15,  # 标题样式列（如果有）
                    'D': 80,  # 内容列
                    'E': 10,  # 信息密度评分
                    'F': 10,  # 信息质量评分
                    'G': 10,  # 是否值得处理
                }
                
                if include_reason:
                    column_widths['H'] = 40  # 评估理由
                
                if suggest_qa_count:
                    # 如果有评估理由，建议问答对数量在I列，否则在H列
                    col_letter = 'I' if include_reason else 'H'
                    column_widths[col_letter] = 15  # 建议问答对数量
                
                for col_letter, width in column_widths.items():
                    if col_letter in worksheet.column_dimensions:
                        worksheet.column_dimensions[col_letter].width = width
                
                # 设置自动换行
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = worksheet.cell(1, 1).alignment.copy(wrapText=True)
        
        print(f"已成功评估内容质量并保存到Excel文件: {output_excel}")
        print(f"总计 {total_rows} 个内容，其中 {len(filtered_df)} 个通过质量筛选")
        
        # 创建仅包含筛选后内容的Excel文件，用于后续问答生成
        filtered_output = f"{os.path.splitext(output_excel)[0]}_filtered{os.path.splitext(output_excel)[1]}"
        
        # 确定要保存的列
        columns_to_save = ['标题', '内容']
        if suggest_qa_count:
            columns_to_save.append('建议问答对数量')
        
        filtered_df[columns_to_save].to_excel(filtered_output, index=False)
        print(f"已将筛选后的内容保存到: {filtered_output}")
        
        return True
            
    except Exception as e:
        print(f"处理Excel文件时出错: {str(e)}")
        return False

if __name__ == "__main__":
    # 设置输入文件路径
    input_excel = "标题内容提取结果.xlsx"
    
    # 处理Excel文件并评估内容质量
    process_excel_and_evaluate(input_excel, min_density_score=5, min_quality_score=5) 