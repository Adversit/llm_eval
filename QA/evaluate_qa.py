import pandas as pd
import json
import requests
import os
from openpyxl import load_workbook
from QA.config import API_KEY, MODEL, API_BASE_URL, DEFAULT_TIMEOUT, DEFAULT_TEMPERATURE

def evaluate_answer_quality(question, answer, content):
    """
    评估问答对的质量，特别是验证答案是否基于原文内容

    参数:
        question (str): 问题
        answer (str): 答案
        content (str): 原始内容

    返回:
        dict: 包含评估结果的字典
    """
    try:
        # 准备提示词
        prompt = f"""
请评估以下问答对的质量，特别是验证答案是否完全基于原文内容。

原文内容:
{content}

问题: {question}

答案: {answer}

评估要求：
1. 从答案中提取关键信息点
2. 检查每个信息点是否能在原文中找到对应依据
3. 给出一个"事实依据分数"（1-10分），表示答案中的信息有多少是来自原文
   - 10分：答案中的所有信息都有明确的原文依据
   - 7-9分：大部分信息有原文依据，但有少量细节可能是推断的
   - 4-6分：约一半信息有原文依据，一半是模型添加的
   - 1-3分：很少信息有原文依据，大部分是模型编造的
4. 给出一个"回答完整性分数"（1-10分），表示答案是否完整回答了问题
5. 给出一个"总体质量分数"（1-10分），综合考虑事实依据和回答完整性
6. 提供评估理由，说明哪些信息有依据，哪些可能是模型编造的

请以JSON格式返回，格式为：
{{
    "factual_score": 事实依据分数(1-10),
    "completeness_score": 回答完整性分数(1-10),
    "overall_score": 总体质量分数(1-10),
    "key_points": ["关键点1", "关键点2", ...],
    "supported_points": ["有依据的关键点1", ...],
    "unsupported_points": ["无依据的关键点1", ...],
    "evaluation_reason": "评估理由"
}}
"""
        
        # 调用API
        url = f"{API_BASE_URL}/v1/chat/completions"

        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": MODEL,
            "messages": [
                {
                    "role": "system",
                    "content": "你是一个专业的问答质量评估专家，擅长判断答案是否基于给定的内容。请确保评估客观公正，并给出合理的分数和判断。"
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "stream": False,
            "max_tokens": 2000,
            "temperature": DEFAULT_TEMPERATURE,
            "response_format": {"type": "json_object"}
        }

        # 发送请求
        response = requests.request("POST", url, json=payload, headers=headers, timeout=DEFAULT_TIMEOUT)
        
        if response.status_code == 200:
            response_data = response.json()
            api_response_content = response_data['choices'][0]['message']['content']

            # 解析JSON响应
            try:
                result = json.loads(api_response_content)
                
                # 确保结果包含预期的字段
                expected_fields = [
                    'factual_score', 'completeness_score', 'overall_score', 
                    'key_points', 'supported_points', 'unsupported_points', 
                    'evaluation_reason'
                ]
                
                for field in expected_fields:
                    if field not in result:
                        if field in ['factual_score', 'completeness_score', 'overall_score']:
                            result[field] = 0
                        elif field in ['key_points', 'supported_points', 'unsupported_points']:
                            result[field] = []
                        else:
                            result[field] = ""
                
                return {
                    'success': True,
                    'evaluation': result
                }
                
            except json.JSONDecodeError:
                return {
                    'success': False,
                    'error': '无法解析API响应为JSON',
                    'raw_response': api_response_content,
                    'evaluation': {
                        'factual_score': 0,
                        'completeness_score': 0,
                        'overall_score': 0,
                        'key_points': [],
                        'supported_points': [],
                        'unsupported_points': [],
                        'evaluation_reason': '评估失败'
                    }
                }
        else:
            return {
                'success': False,
                'error': f'API调用失败: {response.status_code} - {response.text}',
                'evaluation': {
                    'factual_score': 0,
                    'completeness_score': 0,
                    'overall_score': 0,
                    'key_points': [],
                    'supported_points': [],
                    'unsupported_points': [],
                    'evaluation_reason': 'API调用失败'
                }
            }

    except requests.Timeout:
        return {
            'success': False,
            'error': f'API请求超时（超过{DEFAULT_TIMEOUT}秒）',
            'evaluation': {
                'factual_score': 0,
                'completeness_score': 0,
                'overall_score': 0,
                'key_points': [],
                'supported_points': [],
                'unsupported_points': [],
                'evaluation_reason': 'API请求超时'
            }
        }
    except requests.RequestException as e:
        return {
            'success': False,
            'error': f'网络请求错误: {str(e)}',
            'evaluation': {
                'factual_score': 0,
                'completeness_score': 0,
                'overall_score': 0,
                'key_points': [],
                'supported_points': [],
                'unsupported_points': [],
                'evaluation_reason': f'网络请求错误: {str(e)}'
            }
        }
    except Exception as e:
        return {
            'success': False,
            'error': f'评估答案质量时出错: {str(e)}',
            'evaluation': {
                'factual_score': 0,
                'completeness_score': 0,
                'overall_score': 0,
                'key_points': [],
                'supported_points': [],
                'unsupported_points': [],
                'evaluation_reason': f'处理错误: {str(e)}'
            }
        }

def process_qa_and_evaluate(
    qa_excel,
    output_excel=None,
    min_factual_score=7,
    min_overall_score=7,
    sample_percentage=100,
    progress_callback=None
):
    """
    处理问答Excel文件并评估每个问答对的质量
    
    参数:
        qa_excel (str): 问答Excel文件路径
        output_excel (str, optional): 输出Excel文件路径
        min_factual_score (int): 最低事实依据分数阈值，默认为7
        min_overall_score (int): 最低总体质量分数阈值，默认为7
        sample_percentage (float): 抽查的百分比，范围1-100，默认为100（全部评估）
    
    返回:
        bool: 操作是否成功
    """
    try:
        def update_progress(step_name, progress, message=None):
            if progress_callback:
                progress_callback(step_name, progress, message)

        # 设置默认输出路径
        if output_excel is None:
            base_name, ext = os.path.splitext(qa_excel)
            output_excel = f"{base_name}_evaluated{ext}"

        # 确保sample_percentage在有效范围内
        sample_percentage = max(1, min(100, sample_percentage))

        print(f"处理问答Excel文件: {qa_excel}")
        print(f"最低事实依据分数阈值: {min_factual_score}")
        print(f"最低总体质量分数阈值: {min_overall_score}")
        print(f"抽查百分比: {sample_percentage}%")

        update_progress("初始化", 0.05, "开始读取问答Excel文件")

        # 读取Excel文件
        df = pd.read_excel(qa_excel)
        
        # 确保必要的列存在
        required_columns = ['问题', '答案', '内容']
        if not all(col in df.columns for col in required_columns):
            error_msg = f"错误: Excel文件必须包含以下列: {', '.join(required_columns)}"
            print(error_msg)
            update_progress("错误", 0, error_msg)
            return False

        # 添加评估结果列
        df['事实依据分数'] = 0
        df['回答完整性分数'] = 0
        df['总体质量分数'] = 0
        df['关键信息点'] = ''
        df['有依据的信息点'] = ''
        df['无依据的信息点'] = ''
        df['评估理由'] = ''
        df['是否通过质量检验'] = False
        df['是否已评估'] = False
        
        # 计算要评估的行数
        total_rows = len(df)
        sample_size = max(1, int(total_rows * sample_percentage / 100))

        # 随机选择要评估的行索引
        if sample_percentage < 100:
            sample_indices = df.sample(n=sample_size).index.tolist()
            print(f"将随机评估 {sample_size}/{total_rows} 个问答对 ({sample_percentage}%)")
        else:
            sample_indices = df.index.tolist()
            print(f"将评估全部 {total_rows} 个问答对")

        update_progress(
            "准备评估",
            0.1,
            f"将评估 {sample_size}/{total_rows} 个问答对"
        )

        # 遍历选中的问答对
        evaluated_count = 0
        for index in sample_indices:
            row = df.loc[index]
            question = row['问题']
            answer = row['答案']
            content = row['内容']

            evaluated_count += 1
            progress = 0.1 + (evaluated_count / sample_size) * 0.8
            short_question = question[:50] if isinstance(question, str) else ''
            msg = f"正在评估 [{evaluated_count}/{sample_size}]: {short_question}..."
            print(msg)
            update_progress("问答对质量评估", progress, msg)
            
            # 如果答案或内容为空，则跳过
            if pd.isna(answer) or answer.strip() == '' or pd.isna(content) or content.strip() == '':
                skip_msg = "答案或内容为空，跳过评估"
                print(f"  {skip_msg}")
                df.at[index, '评估理由'] = '答案或内容为空'
                df.at[index, '是否已评估'] = True
                update_progress("问答对质量评估", progress, f"{msg} {skip_msg}")
                continue
                
            # 评估答案质量
            eval_result = evaluate_answer_quality(question, answer, content)
            
            if eval_result['success']:
                evaluation = eval_result['evaluation']
                
                # 更新DataFrame
                df.at[index, '事实依据分数'] = evaluation['factual_score']
                df.at[index, '回答完整性分数'] = evaluation['completeness_score']
                df.at[index, '总体质量分数'] = evaluation['overall_score']
                df.at[index, '关键信息点'] = ', '.join(evaluation['key_points'])
                df.at[index, '有依据的信息点'] = ', '.join(evaluation['supported_points'])
                df.at[index, '无依据的信息点'] = ', '.join(evaluation['unsupported_points'])
                df.at[index, '评估理由'] = evaluation['evaluation_reason']
                df.at[index, '是否已评估'] = True
                
                # 判断是否通过质量检验
                passed = (evaluation['factual_score'] >= min_factual_score and 
                          evaluation['overall_score'] >= min_overall_score)
                df.at[index, '是否通过质量检验'] = passed
                
                status = "通过" if passed else "未通过"
                detail_msg = (
                    f"评估结果: 事实依据={evaluation['factual_score']}, "
                    f"完整性={evaluation['completeness_score']}, "
                    f"总体质量={evaluation['overall_score']}, 状态={status}"
                )
                print(f"  {detail_msg}")
                update_progress("问答对质量评估", progress, detail_msg)
            else:
                err_msg = f"评估失败: {eval_result.get('error', '未知错误')}"
                print(f"  {err_msg}")
                df.at[index, '评估理由'] = f"评估失败: {eval_result.get('error', '未知错误')}"
                df.at[index, '是否已评估'] = True
                update_progress("问答对质量评估", progress, err_msg)
        
        # 创建一个筛选后的DataFrame（只包含已评估且通过质量检验的行）
        filtered_df = df[
            (df['是否已评估'] == True) & 
            (df['事实依据分数'] >= min_factual_score) & 
            (df['总体质量分数'] >= min_overall_score) & 
            (df['是否通过质量检验'] == True)
        ].copy()
        
        update_progress("保存结果", 0.95, "正在保存评估结果")

        # 保存完整评估结果
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='全部评估结果', index=False)
            filtered_df.to_excel(writer, sheet_name='筛选后结果', index=False)
            
            # 设置列宽
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                column_widths = {
                    'A': 40,  # 标题
                    'B': 10,  # 问题编号
                    'C': 50,  # 问题
                    'D': 80,  # 答案
                    'E': 80,  # 内容
                    'F': 10,  # 事实依据分数
                    'G': 10,  # 回答完整性分数
                    'H': 10,  # 总体质量分数
                    'I': 40,  # 关键信息点
                    'J': 40,  # 有依据的信息点
                    'K': 40,  # 无依据的信息点
                    'L': 60,  # 评估理由
                    'M': 10,  # 是否通过质量检验
                    'N': 10,  # 是否已评估
                }
                
                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
                
                # 设置自动换行
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = worksheet.cell(1, 1).alignment.copy(wrapText=True)
        
        # 创建仅包含筛选后内容的Excel文件
        filtered_output = f"{os.path.splitext(output_excel)[0]}_filtered{os.path.splitext(output_excel)[1]}"
        filtered_df.to_excel(filtered_output, index=False)
        
        # 计算评估统计信息
        evaluated_total = df['是否已评估'].sum()
        passed_total = df['是否通过质量检验'].sum()
        pass_rate = (passed_total / evaluated_total * 100) if evaluated_total > 0 else 0
        
        summary_msg = (
            f"总计 {total_rows} 个问答对，评估了 {evaluated_total} 个 "
            f"({evaluated_total/total_rows*100:.1f}%)，其中 {passed_total} 个通过质量筛选 "
            f"(通过率: {pass_rate:.1f}%)"
        )
        print(f"已成功评估问答对质量并保存到Excel文件: {output_excel}")
        print(summary_msg)
        print(f"已将筛选后的问答对保存到: {filtered_output}")
        update_progress("完成", 1.0, summary_msg)

        return True
            
    except Exception as e:
        error_msg = f"处理Excel文件时出错: {str(e)}"
        print(error_msg)
        update_progress("错误", 0, error_msg)
        return False

if __name__ == "__main__":
    # 设置输入文件路径
    qa_excel = "标题内容提取结果_evaluated_filtered_QA.xlsx"
    
    # 处理问答Excel文件并评估质量
    process_qa_and_evaluate(qa_excel, min_factual_score=7, min_overall_score=7) 
